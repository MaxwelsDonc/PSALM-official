#!/usr/bin/env python3
"""
RQ3 Analysis Script: PSALM vs ART/MT-ART Effectiveness Comparison

This script generates two Excel reports comparing PSALM effectiveness against:
- RQ3_phase1.xlsx: PSALM vs ART (source test case selection phase)
- RQ3_phase2.xlsx: PSALM vs MT-ART (MG selection phase)

Each project generates one worksheet with specified sheet name mapping and order.
"""

import os
import sys
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import wilcoxon
import yaml


def load_config(config_path):
    """Load configuration from YAML file."""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_json_data(file_path):
    """Load JSON data from file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: File not found: {file_path}")
        return None
    except json.JSONDecodeError as e:
        print(f"Warning: JSON decode error in {file_path}: {e}")
        return None


def vargha_delaney_a12(x, y):
    """
    计算Vargha-Delaney A12效应量

    Args:
        x: 第一组数据 (list or array)
        y: 第二组数据 (list or array)

    Returns:
        float: A12效应量 (0-1之间)
    """
    x = np.array(x)
    y = np.array(y)

    # 计算所有配对比较
    m = len(x)
    n = len(y)

    if m == 0 or n == 0:
        return 0.5

    # 计算x中每个值大于y中每个值的次数
    wins = 0
    ties = 0

    for xi in x:
        for yi in y:
            if xi > yi:
                wins += 1
            elif xi == yi:
                ties += 1

    # A12 = (wins + 0.5 * ties) / (m * n)
    a12 = (wins + 0.5 * ties) / (m * n)
    return a12


# Constants
RUNS = 30

# Sheet name mapping (same as RQ2)
SHEET_NAME_MAP = {
    "jackson_project": "ParseI",
    "jfreeChart_project": "CLine",
    "lang_project": "IsDay",
    "math1_project": "Conv",
    "math2_project": "CopyS",
    "geometricSum_project": "GeS",
    "incomeTax_project": "InT",
    "mortgageRate_project": "MoR",
}

# Sheet order (same as RQ2)
SHEET_ORDER = ["MoR", "GeS", "InT", "CopyS", "ParseI", "CLine", "Conv", "IsDay"]


def map_sheet_name(project_name):
    """Map project name to sheet name."""
    return SHEET_NAME_MAP.get(project_name, project_name)


def analyze_project(
    project_name,
    project_path,
    base_dir,
    phase,
    baseline_method,
    mutants_list,
    partition_num,
):
    """
    Analyze a single project comparing PSALM vs baseline method.

    Args:
        project_name: Name of the project
        project_path: Path to the project from config
        base_dir: Base directory path
        phase: 'phase1' or 'phase2'
        baseline_method: 'ART' or 'MTART'
        mutants_list: List of mutants to analyze from config
        partition_num: Number of partitions from config

    Returns:
        List of result dictionaries for each mutant plus summary
    """
    # Construct data paths using the project path from config
    project_root = os.path.join(base_dir, project_path, "raw_results")
    phase_dir = os.path.join(project_root, phase)
    psalm_file = os.path.join(phase_dir, "P-measure_partition.json")

    if baseline_method == "ART":
        baseline_file = os.path.join(phase_dir, "P-measure_art.json")
        baseline_label = "ART"
    else:  # MTART
        baseline_file = os.path.join(phase_dir, "P-measure_mtart.json")
        baseline_label = "MT-ART"

    # Load data
    psalm_data = load_json_data(psalm_file)
    baseline_data = load_json_data(baseline_file)

    if not psalm_data or not baseline_data:
        print(f"Warning: Missing data files for {project_name} {phase}")
        return []

    # Find common mutants and filter by config
    common_mutants = (
        set(psalm_data.keys()) & set(baseline_data.keys()) & set(mutants_list)
    )
    if not common_mutants:
        print(
            f"Warning: No common mutants found for {project_name} {phase} (config specifies {len(mutants_list)} mutants)"
        )
        return []

    results = []
    psalm_runs_matrix = []
    baseline_runs_matrix = []

    # Process each mutant
    for mutant in sorted(common_mutants):
        psalm_mutant_data = psalm_data[mutant]
        baseline_mutant_data = baseline_data[mutant]

        # Find common test case numbers and filter by 3*partition_num
        target_tc_num = str(3 * partition_num)
        common_tcs = set(psalm_mutant_data.keys()) & set(baseline_mutant_data.keys())

        # Check if target test case number exists
        if target_tc_num not in common_tcs:
            print(
                f"Warning: Test case number {target_tc_num} not found for {mutant} in {project_name}"
            )
            continue

        # Use the target test case number (3 * partition_num)
        tc_num = target_tc_num

        psalm_runs = psalm_mutant_data[tc_num][:RUNS]
        baseline_runs = baseline_mutant_data[tc_num][:RUNS]

        if len(psalm_runs) < RUNS or len(baseline_runs) < RUNS:
            continue

        # Calculate statistics
        mean_psalm = np.mean(psalm_runs)
        mean_baseline = np.mean(baseline_runs)

        # Improvement percentage
        if mean_baseline != 0:
            improvement = (mean_psalm - mean_baseline) / mean_baseline * 100
        else:
            improvement = 0.0

        # Wilcoxon signed-rank test (two-tailed)
        try:
            _, p_value = wilcoxon(
                psalm_runs, baseline_runs, zero_method="pratt", alternative="two-sided"
            )
        except:
            p_value = np.nan

        # Vargha-Delaney A12 effect size
        a12 = vargha_delaney_a12(psalm_runs, baseline_runs)

        # Store result
        result = {
            "mutant": mutant,
            "PSALM": mean_psalm,
            baseline_label: mean_baseline,
            "Improvement(%)": improvement,
            "p(PSALMvsBaseline)": p_value,
            "A12(PSALMvsBaseline)": a12,
        }
        results.append(result)

        # Collect runs for summary calculation
        psalm_runs_matrix.append(psalm_runs)
        baseline_runs_matrix.append(baseline_runs)

    # Calculate summary row based on 30 run-level averages
    if psalm_runs_matrix and baseline_runs_matrix:
        psalm_matrix = np.array(psalm_runs_matrix)  # Shape: (mutants, 30)
        baseline_matrix = np.array(baseline_runs_matrix)  # Shape: (mutants, 30)

        # Calculate column-wise means (30 run-level averages)
        psalm_run_averages = np.mean(psalm_matrix, axis=0)  # Shape: (30,)
        baseline_run_averages = np.mean(baseline_matrix, axis=0)  # Shape: (30,)

        # Summary statistics
        summary_psalm_mean = np.mean(psalm_run_averages)
        summary_baseline_mean = np.mean(baseline_run_averages)

        if summary_baseline_mean != 0:
            summary_improvement = (
                (summary_psalm_mean - summary_baseline_mean)
                / summary_baseline_mean
                * 100
            )
        else:
            summary_improvement = 0.0

        # Wilcoxon test on run-level averages
        psalm_run_averages = [round(x, 4) for x in psalm_run_averages]
        baseline_run_averages = [round(x, 4) for x in baseline_run_averages]
        try:
            _, summary_p_value = wilcoxon(
                psalm_run_averages,
                baseline_run_averages,
                zero_method="pratt",
                alternative="two-sided",
            )
        except:
            summary_p_value = np.nan

        # A12 on run-level averages
        summary_a12 = vargha_delaney_a12(psalm_run_averages, baseline_run_averages)

        # Add summary row
        summary = {
            "mutant": "SUMMARY",
            "PSALM": summary_psalm_mean,
            baseline_label: summary_baseline_mean,
            "Improvement(%)": summary_improvement,
            "p(PSALMvsBaseline)": summary_p_value,
            "A12(PSALMvsBaseline)": summary_a12,
        }
        results.append(summary)

    return results


def create_excel_report(all_results, output_file, baseline_label):
    """
    Create Excel report with formatted sheets.

    Args:
        all_results: Dictionary mapping project names to result lists
        output_file: Output Excel file path
        baseline_label: 'ART' or 'MT-ART'
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Define column names
    columns = [
        "mutant",
        "PSALM",
        baseline_label,
        "Improvement(%)",
        "p(PSALMvsBaseline)",
        "A12(PSALMvsBaseline)",
    ]

    # Sort projects by sheet order
    project_names = list(all_results.keys())
    sheet_names = [map_sheet_name(proj) for proj in project_names]

    # Create mapping and sort by order
    sheet_project_pairs = list(zip(sheet_names, project_names))
    sheet_project_pairs.sort(
        key=lambda x: (
            SHEET_ORDER.index(x[0]) if x[0] in SHEET_ORDER else len(SHEET_ORDER)
        )
    )

    for sheet_name, project_name in sheet_project_pairs:
        results = all_results[project_name]
        if not results:
            continue

        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)

        # Create DataFrame
        df = pd.DataFrame(results, columns=columns)

        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Format summary row (last row)
        if len(results) > 0:
            summary_row = len(results) + 1  # +1 for header
            for cell in ws[summary_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_file)
    print(f"Excel report saved: {output_file}")


def main():
    """Main function to orchestrate the analysis."""
    # Get script directory and resolve paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)

    print(f"Base directory: {base_dir}")

    # Load configuration
    config_file = os.path.join(script_dir, "config.yaml")
    print(f"Config file: {config_file}")

    config = load_config(config_file)
    if not config:
        print("Error: Failed to load configuration")
        return

    projects = config.get("projects", [])
    if not projects:
        print("Error: No projects found in configuration")
        return

    # Output files
    phase1_output = os.path.join(script_dir, "RQ3_phase1.xlsx")
    phase2_output = os.path.join(script_dir, "RQ3_phase2.xlsx")

    print(f"Phase1 output file: {phase1_output}")
    print(f"Phase2 output file: {phase2_output}")

    # Analyze Phase 1: PSALM vs ART
    print("\n=== Phase 1 Analysis: PSALM vs ART ===")
    phase1_results = {}
    total_rows_phase1 = 0

    for project in projects:
        project_name = project["name"]
        project_path = project["path"]
        mutants_list = project["mutants"]
        partition_num = project["partition_num"]
        print(f"Analyzing {project_name} (Phase 1)...")

        results = analyze_project(
            project_name,
            project_path,
            base_dir,
            "phase1",
            "ART",
            mutants_list,
            partition_num,
        )
        if results:
            phase1_results[project_name] = results
            total_rows_phase1 += len(results)
            print(f"Completed analysis for {project_name}: {len(results)} rows")
        else:
            print(f"No results for {project_name} (Phase 1)")

    # Create Phase 1 Excel report
    if phase1_results:
        print(f"\nCreating Phase 1 Excel report: {phase1_output}")
        create_excel_report(phase1_results, phase1_output, "ART")

    # Analyze Phase 2: PSALM vs MT-ART
    print("\n=== Phase 2 Analysis: PSALM vs MT-ART ===")
    phase2_results = {}
    total_rows_phase2 = 0

    for project in projects:
        project_name = project["name"]
        project_path = project["path"]
        mutants_list = project["mutants"]
        partition_num = project["partition_num"]
        print(f"Analyzing {project_name} (Phase 2)...")

        results = analyze_project(
            project_name,
            project_path,
            base_dir,
            "phase2",
            "MTART",
            mutants_list,
            partition_num,
        )
        if results:
            phase2_results[project_name] = results
            total_rows_phase2 += len(results)
            print(f"Completed analysis for {project_name}: {len(results)} rows")
        else:
            print(f"No results for {project_name} (Phase 2)")

    # Create Phase 2 Excel report
    if phase2_results:
        print(f"\nCreating Phase 2 Excel report: {phase2_output}")
        create_excel_report(phase2_results, phase2_output, "MT-ART")

    # Summary
    print(f"\n=== Analysis Summary ===")
    print(f"Phase 1 - Projects analyzed: {len(phase1_results)}")
    print(f"Phase 1 - Total rows generated: {total_rows_phase1}")
    print(f"Phase 2 - Projects analyzed: {len(phase2_results)}")
    print(f"Phase 2 - Total rows generated: {total_rows_phase2}")

    if phase1_results:
        print(f"Phase 1 results saved to: RQ3_phase1.xlsx")
    if phase2_results:
        print(f"Phase 2 results saved to: RQ3_phase2.xlsx")


if __name__ == "__main__":
    main()
