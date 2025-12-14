#!/usr/bin/env python3
"""
RQ1 Analysis Script
分析不同测试策略的效果，生成Excel报告
"""

import json
from pathlib import Path
from typing import Final

import numpy as np
from openpyxl.styles import Font, PatternFill
import pandas as pd
from scipy.stats import wilcoxon
import yaml

RUNS: Final[int] = 30  # 默认


def vargha_delaney_a12(x, y):
    """
    计算Vargha-Delaney A12效应量

    Args:
        x: 第一组数据 (list or array)
        y: 第二组数据 (list or array)

    Returns:
        float: A12效应量 (0-1之间)
    """
    x = np.array(x)
    y = np.array(y)

    # 计算所有配对比较
    m = len(x)
    n = len(y)

    if m == 0 or n == 0:
        return 0.5

    # 计算x中每个值大于y中每个值的次数
    wins = 0
    ties = 0

    for xi in x:
        for yi in y:
            if xi > yi:
                wins += 1
            elif xi == yi:
                ties += 1

    # A12 = (wins + 0.5 * ties) / (m * n)
    a12 = (wins + 0.5 * ties) / (m * n)
    return a12


def load_config(config_path):
    """加载配置文件"""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_json_data(file_path):
    """加载JSON数据文件"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Warning: File not found: {file_path}")
        return {}
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON in file: {file_path}")
        return {}


def get_project_data_paths(project_path, base_dir):
    """获取项目数据文件路径"""
    project_full_path = Path(base_dir) / project_path

    # 检查是否为Python项目
    if "python" in project_path:
        raw_results_path = project_full_path / "raw_results"
    else:
        # Java项目路径结构可能不同，这里假设类似的结构
        raw_results_path = project_full_path / "raw_results"

    phase1_path = raw_results_path / "phase1"
    phase2_path = raw_results_path / "phase2"

    return {
        "phase1": {
            "partition": phase1_path / "P-measure_partition.json",
            "random": phase1_path / "P-measure_random.json",
            "art": phase1_path / "P-measure_art.json",
        },
        "phase2": {
            "partition": phase2_path / "P-measure_partition.json",
            "random": phase2_path / "P-measure_random.json",
            "mt-art": phase2_path / "P-measure_mtart.json",
        },
    }


def analyze_project(project_config, base_dir):
    """分析单个项目 (只比较 partition vs random，不再统计 art/mtart)"""
    project_name = project_config["name"]
    project_path = project_config["path"]
    partition_num = project_config["partition_num"]
    mutants = project_config["mutants"]

    # 目标测试用例数
    target_test_cases = str(partition_num * 3)

    print(f"\nAnalyzing project: {project_name}")
    print(f"Target test cases: {target_test_cases}")
    print(f"Mutants to analyze: {len(mutants)}")

    # 获取路径
    project_root = Path(base_dir) / project_path / "raw_results"
    phase_paths = {
        "phase1": {
            "partition": project_root / "phase1" / "P-measure_partition.json",
            "random": project_root / "phase1" / "P-measure_random.json",
        },
        "phase2": {
            "partition": project_root / "phase2" / "P-measure_partition.json",
            "random": project_root / "phase2" / "P-measure_random.json",
        },
    }

    results = []

    for phase in ["phase1", "phase2"]:
        print(f"  Processing {phase}...")

        # 读数据
        partition_data = load_json_data(phase_paths[phase]["partition"])
        random_data = load_json_data(phase_paths[phase]["random"])

        # 存所有 mutants 的均值
        all_partition_means = []
        all_random_means = []

        # 收集 per-run 矩阵，用于 summary
        per_run_partition = []
        per_run_random = []

        for mutant in mutants:
            if (
                mutant in partition_data
                and target_test_cases in partition_data[mutant]
                and mutant in random_data
                and target_test_cases in random_data[mutant]
            ):

                part_vals = partition_data[mutant][target_test_cases]
                rand_vals = random_data[mutant][target_test_cases]
                # 只取前 30 个，要求两边都至少有 30 个
                if len(part_vals) < RUNS or len(rand_vals) < RUNS:
                    continue
                part_vals = part_vals[:RUNS]
                rand_vals = rand_vals[:RUNS]

                if len(part_vals) == len(rand_vals) and len(part_vals) > 0:
                    # mutant 均值
                    part_mean = np.mean(part_vals)
                    rand_mean = np.mean(rand_vals)

                    # Wilcoxon + A12
                    try:
                        _, p_val = wilcoxon(
                            part_vals, rand_vals, alternative="two-sided"
                        )
                        a12 = vargha_delaney_a12(part_vals, rand_vals)
                    except Exception as e:
                        print(f"    Warning: Wilcoxon failed for {mutant}: {e}")
                        p_val, a12 = np.nan, np.nan

                    results.append(
                        {
                            "mutant": mutant,
                            "phase": phase,
                            "partition": part_mean,
                            "random": rand_mean,
                            "p-value(partition vs random)": p_val,
                            "A12(partition vs random)": a12,
                        }
                    )

                    all_partition_means.append(part_mean)
                    all_random_means.append(rand_mean)
                    per_run_partition.append(part_vals)
                    per_run_random.append(rand_vals)

        # === SUMMARY 行 ===
        if per_run_partition and per_run_random:
            # 转成矩阵 [mutant, runs]
            mat_part = np.array(per_run_partition)  # shape: [M, 50]
            mat_rand = np.array(per_run_random)  # shape: [M, 50]

            # 每一轮运行的平均 (长度=50)
            run_part = mat_part.mean(axis=0)
            run_rand = mat_rand.mean(axis=0)

            try:
                _, summary_p = wilcoxon(run_part, run_rand, alternative="two-sided")
                summary_a12 = vargha_delaney_a12(run_part, run_rand)
            except Exception as e:
                print(f"    Warning: Wilcoxon failed for SUMMARY: {e}")
                summary_p, summary_a12 = np.nan, np.nan

            results.append(
                {
                    "mutant": "SUMMARY",
                    "phase": phase,
                    "partition": float(np.mean(all_partition_means)),
                    "random": float(np.mean(all_random_means)),
                    "p-value(partition vs random)": summary_p,
                    "A12(partition vs random)": summary_a12,
                }
            )

    return results


def create_excel_report(all_results, output_path):
    """创建Excel报告"""
    print(f"Creating Excel report: {output_path}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wrote_any = False  # 标记是否写过sheet

        for project_name, project_results in all_results.items():
            if not project_results:
                continue

            columns = [
                "mutant",
                "phase",
                "partition",
                "random",
                "p-value(partition vs random)",
                "A12(partition vs random)",
            ]

            # 创建DataFrame
            df = pd.DataFrame(project_results)[columns]

            # 写入sheet
            df.to_excel(writer, sheet_name=project_name, index=False)
            wrote_any = True

            # 获取工作表进行格式化
            worksheet = writer.sheets[project_name]

            # 设置列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 高亮SUMMARY行
            summary_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            summary_font = Font(bold=True)

            for row in worksheet.iter_rows():
                if row[0].value == "SUMMARY":
                    for cell in row:
                        cell.fill = summary_fill
                        cell.font = summary_font

        # 如果没有任何结果，写一个占位sheet，避免openpyxl报错
        if not wrote_any:
            empty_df = pd.DataFrame([{"Info": "No results available"}])
            empty_df.to_excel(writer, sheet_name="Empty", index=False)


def main():
    """主函数"""
    # 设置路径
    base_dir = Path(__file__).parent.parent
    config_path = Path(__file__).parent / "config.yaml"
    output_path = Path(__file__).parent / "RQ1.xlsx"

    print(f"Base directory: {base_dir}")
    print(f"Config file: {config_path}")
    print(f"Output file: {output_path}")

    # 加载配置
    config = load_config(config_path)

    # 分析所有项目
    all_results = {}

    for project_config in config["projects"]:
        project_name = project_config["name"]
        try:
            project_results = analyze_project(project_config, base_dir)
            all_results[project_name] = project_results
            print(f"Completed analysis for {project_name}: {len(project_results)} rows")
        except Exception as e:
            print(f"Error analyzing project {project_name}: {e}")
            all_results[project_name] = []

    # 创建Excel报告
    create_excel_report(all_results, output_path)

    print(f"\nAnalysis completed! Results saved to: {output_path}")

    # 打印统计信息
    total_rows = sum(len(results) for results in all_results.values())
    print(f"Total rows generated: {total_rows}")
    print(f"Projects analyzed: {len(all_results)}")


if __name__ == "__main__":
    main()
