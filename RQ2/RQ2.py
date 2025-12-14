#!/usr/bin/env python3
"""
RQ2 Analysis Script
基于与 RQ1 相同的实验设置，生成 Excel 报告比较 PSALM 在两个测试阶段的有效性：
1) 选择源测试用例（phase1）
2) 选择 MGs（phase2）

输出：RQ2/RQ2.xlsx，每个项目一个工作表，包含每个 mutant 的对比及 SUMMARY 行。
"""

from pathlib import Path
from typing import Final, List, Dict

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from scipy.stats import wilcoxon

# 复用 RQ1 中的辅助函数
from RQ1.RQ1 import load_config, load_json_data, vargha_delaney_a12


# 固定运行次数
RUNS: Final[int] = 30


# 工作表名映射
SHEET_NAME_MAP = {
    "jackson_project": "ParseI",
    "jfreeChart_project": "CLine",
    "lang_project": "IsDay",
    "math1_project": "Conv",
    "math2_project": "CopyS",
    "geometricSum_project": "GeS",
    "incomeTax_project": "InT",
    "mortgageRate_project": "MoR",
}


def map_sheet_name(project_name: str) -> str:
    """将项目名映射为指定工作表名，若未映射则保留原名。"""
    return SHEET_NAME_MAP.get(project_name, project_name)


# 自定义工作表顺序（若存在则按此顺序写入）
SHEET_ORDER = [
    "MoR",
    "GeS",
    "InT",
    "CopyS",
    "ParseI",
    "CLine",
    "Conv",
    "IsDay",
]


def analyze_project(project_config: Dict, base_dir: Path) -> List[Dict]:
    """分析单个项目，比较 phase1 与 phase2 的 PSALM（partition）结果。

    返回每个 mutant 的一行结果，以及最后的 SUMMARY 行。
    列：mutant, phase1_mean, phase2_mean, Improvement(%), p(MGvsSrc), A12(MGvsSrc)
    """

    project_name = project_config["name"]
    project_path = project_config["path"]
    partition_num = project_config["partition_num"]
    mutants = project_config["mutants"]

    target_test_cases = str(partition_num * 3)

    project_root = Path(base_dir) / project_path / "raw_results"
    phase1_file = project_root / "phase1" / "P-measure_partition.json"
    phase2_file = project_root / "phase2" / "P-measure_partition.json"

    # 读数据
    phase1_data = load_json_data(phase1_file)
    phase2_data = load_json_data(phase2_file)

    results: List[Dict] = []

    # 收集用于 SUMMARY 的数据
    all_phase1_means: List[float] = []
    all_phase2_means: List[float] = []
    per_run_phase1: List[List[float]] = []
    per_run_phase2: List[List[float]] = []

    for mutant in mutants:
        if (
            mutant in phase1_data
            and target_test_cases in phase1_data[mutant]
            and mutant in phase2_data
            and target_test_cases in phase2_data[mutant]
        ):
            vals1 = phase1_data[mutant][target_test_cases]
            vals2 = phase2_data[mutant][target_test_cases]

            # 保证至少有 RUNS 个；只取前 RUNS 个
            if len(vals1) < RUNS or len(vals2) < RUNS:
                continue
            vals1 = vals1[:RUNS]
            vals2 = vals2[:RUNS]

            # 计算均值与统计量
            mean1 = float(np.mean(vals1))
            mean2 = float(np.mean(vals2))
            improvement = (mean2 - mean1) / mean1 * 100 if mean1 != 0 else float("nan")

            try:
                # phase2 vs phase1，two-sided，zero_method='pratt'
                _, p_val = wilcoxon(
                    vals2, vals1, alternative="two-sided", zero_method="pratt"
                )
                a12 = float(vargha_delaney_a12(vals2, vals1))
            except Exception as e:
                print(f"    Warning: Wilcoxon failed for {project_name}:{mutant}: {e}")
                p_val, a12 = float("nan"), float("nan")

            results.append(
                {
                    "mutant": mutant,
                    "MG": mean2,
                    "st": mean1,
                    "Improvement(%)": improvement,
                    "p(MGvsSrc)": p_val,
                    "A12(MGvsSrc)": a12,
                }
            )

            all_phase1_means.append(mean1)
            all_phase2_means.append(mean2)
            per_run_phase1.append(vals1)
            per_run_phase2.append(vals2)

    # === SUMMARY 行（按 RQ1 逻辑）===
    if per_run_phase1 and per_run_phase2:
        mat1 = np.array(per_run_phase1)  # [mutants, RUNS]
        mat2 = np.array(per_run_phase2)  # [mutants, RUNS]

        run1 = mat1.mean(axis=0)  # 30 个运行的列均值
        run2 = mat2.mean(axis=0)

        try:
            _, summary_p = wilcoxon(
                run2, run1, alternative="two-sided", zero_method="pratt"
            )
            summary_a12 = float(vargha_delaney_a12(run2, run1))
        except Exception as e:
            print(f"    Warning: Wilcoxon failed for SUMMARY in {project_name}: {e}")
            summary_p, summary_a12 = float("nan"), float("nan")

        summary_mean1 = float(np.mean(all_phase1_means))
        summary_mean2 = float(np.mean(all_phase2_means))
        summary_improvement = (
            (summary_mean2 - summary_mean1) / summary_mean1 * 100
            if summary_mean1 != 0
            else float("nan")
        )

        results.append(
            {
                "mutant": "SUMMARY",
                "MG": summary_mean2,
                "st": summary_mean1,
                "Improvement(%)": summary_improvement,
                "p(MGvsSrc)": summary_p,
                "A12(MGvsSrc)": summary_a12,
            }
        )

    return results


def create_excel_report(all_results: Dict[str, List[Dict]], output_path: Path) -> None:
    """创建 RQ2 Excel 报告，每项目一个 sheet，自动列宽，SUMMARY 行高亮。"""
    print(f"Creating Excel report: {output_path}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wrote_any = False

        # 先构建 (project_name, rows, sheet_name) 列表
        items = [
            (project_name, project_rows, map_sheet_name(project_name))
            for project_name, project_rows in all_results.items()
        ]

        # 按自定义顺序排序；未在顺序中的放最后并按名称排序
        def sort_key(item):
            _, _, sheet_name = item
            try:
                idx = SHEET_ORDER.index(sheet_name)
            except ValueError:
                idx = len(SHEET_ORDER) + 1
            return (idx, sheet_name)

        items.sort(key=sort_key)

        for project_name, project_rows, sheet_name in items:
            if not project_rows:
                continue

            columns = [
                "mutant",
                "MG",
                "st",
                "Improvement(%)",
                "p(MGvsSrc)",
                "A12(MGvsSrc)",
            ]

            df = pd.DataFrame(project_rows)[columns]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            wrote_any = True

            ws = writer.sheets[sheet_name]

            # 自动列宽（同 RQ1）
            for column in ws.columns:
                max_len = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 22)

            # 高亮 SUMMARY 行
            summary_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            summary_font = Font(bold=True)
            for row in ws.iter_rows():
                if row[0].value == "SUMMARY":
                    for cell in row:
                        cell.fill = summary_fill
                        cell.font = summary_font

        if not wrote_any:
            pd.DataFrame([{"Info": "No results available"}]).to_excel(
                writer, sheet_name="Empty", index=False
            )


def main():
    # 路径与配置（与 RQ1 保持一致）
    base_dir = Path(__file__).parent.parent
    config_path = Path(__file__).parent.parent / "RQ2" / "config.yaml"
    output_path = Path(__file__).parent / "RQ2.xlsx"

    print(f"Base directory: {base_dir}")
    print(f"Config file: {config_path}")
    print(f"Output file: {output_path}")

    config = load_config(config_path)

    all_results: Dict[str, List[Dict]] = {}
    projects = config.get("projects", [])

    for project_config in projects:
        project_name = project_config.get("name", "unknown_project")
        try:
            rows = analyze_project(project_config, base_dir)
            all_results[project_name] = rows
            print(f"Completed analysis for {project_name}: {len(rows)} rows")
        except Exception as e:
            print(f"Error analyzing project {project_name}: {e}")
            all_results[project_name] = []

    create_excel_report(all_results, output_path)

    # 输出提示信息
    print("\nAnalysis completed! Results saved to: RQ2.xlsx")
    total_rows = sum(len(v) for v in all_results.values())
    print(f"Projects analyzed: {len(all_results)}")
    print(f"Total rows generated: {total_rows}")


if __name__ == "__main__":
    main()
